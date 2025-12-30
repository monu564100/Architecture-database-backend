import os
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
from typing import Optional, List, Dict, Any
import logging
import json

logger = logging.getLogger(__name__)

class ExcelService:
    """Service for managing prompt-response data in Excel sheets"""
    
    def __init__(self, excel_path: str = "knowledge_base.xlsx"):
        self.excel_path = excel_path
        self._initialize_excel()
    
    def _initialize_excel(self):
        """Create Excel file with required sheets if it doesn't exist"""
        if not os.path.exists(self.excel_path):
            logger.info(f"Creating new Excel file: {self.excel_path}")
            wb = Workbook()
            
            # Create sheets for each category
            categories = ["architecture", "ui", "database", "api", "prompts", "general"]
            
            # Remove default sheet
            default_sheet = wb.active
            default_sheet.title = categories[0]
            
            # Create header for first sheet
            self._add_headers(default_sheet)
            
            # Create other sheets
            for category in categories[1:]:
                ws = wb.create_sheet(title=category)
                self._add_headers(ws)
            
            wb.save(self.excel_path)
            logger.info("Excel file created with all category sheets")
        else:
            logger.info(f"Using existing Excel file: {self.excel_path}")
    
    def _add_headers(self, worksheet):
        """Add headers to a worksheet"""
        headers = [
            "ID", "Timestamp", "Prompt", "Response", "Category", 
            "Prompt_Embedding", "Usage_Count", "Last_Used", "User_Rating"
        ]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
    
    def get_all_data(self, category: str) -> pd.DataFrame:
        """Get all data from a specific category sheet"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name=category)
            return df
        except Exception as e:
            logger.warning(f"Error reading sheet {category}: {e}")
            return pd.DataFrame()
    
    def add_entry(
        self,
        prompt: str,
        response: str,
        category: str,
        embedding: List[float] = None
    ) -> str:
        """Add a new prompt-response pair to the Excel sheet"""
        logger.info(f"💾 Adding entry to Excel sheet: {category}")
        logger.info(f"   Prompt: {prompt[:80]}...")
        logger.info(f"   Response length: {len(response)} chars")
        
        try:
            # Read existing data
            try:
                df = pd.read_excel(self.excel_path, sheet_name=category)
                logger.info(f"   Existing rows in sheet: {len(df)}")
            except Exception as read_err:
                logger.warning(f"   Could not read existing sheet: {read_err}")
                df = pd.DataFrame()
            
            # Generate new ID
            new_id = f"{category[:3].upper()}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{len(df) + 1}"
            logger.info(f"   New entry ID: {new_id}")
            
            # Create new entry
            new_entry = {
                "ID": new_id,
                "Timestamp": datetime.now().isoformat(),
                "Prompt": prompt,
                "Response": response,
                "Category": category,
                "Prompt_Embedding": json.dumps(embedding) if embedding else "",
                "Usage_Count": 1,
                "Last_Used": datetime.now().isoformat(),
                "User_Rating": 0
            }
            
            # Append to dataframe
            new_df = pd.DataFrame([new_entry])
            df = pd.concat([df, new_df], ignore_index=True)
            
            # Save back to Excel
            logger.info(f"   Saving to Excel file: {self.excel_path}")
            with pd.ExcelWriter(self.excel_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=category, index=False)
            
            logger.info(f"✅ Successfully added entry {new_id} to {category} sheet")
            logger.info(f"   Total rows now: {len(df)}")
            return new_id
            
        except Exception as e:
            logger.error(f"❌ Error adding entry: {e}")
            import traceback
            logger.error(traceback.format_exc())
            raise
    
    def update_usage(self, entry_id: str, category: str):
        """Update usage count and last used timestamp for an entry"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name=category)
            
            mask = df["ID"] == entry_id
            if mask.any():
                df.loc[mask, "Usage_Count"] = df.loc[mask, "Usage_Count"] + 1
                df.loc[mask, "Last_Used"] = datetime.now().isoformat()
                
                with pd.ExcelWriter(self.excel_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=category, index=False)
                    
                logger.info(f"Updated usage for {entry_id}")
        except Exception as e:
            logger.error(f"Error updating usage: {e}")
    
    def get_all_prompts_with_embeddings(self, category: str) -> List[Dict[str, Any]]:
        """Get all prompts with their embeddings for similarity matching"""
        try:
            logger.info(f"📖 Reading Excel sheet: {category}")
            df = pd.read_excel(self.excel_path, sheet_name=category)
            
            logger.info(f"   Found {len(df)} rows in {category} sheet")
            
            if df.empty:
                logger.info(f"   Sheet is empty, returning []")
                return []
            
            results = []
            for idx, row in df.iterrows():
                embedding = None
                if pd.notna(row.get("Prompt_Embedding")) and row["Prompt_Embedding"]:
                    try:
                        embedding = json.loads(row["Prompt_Embedding"])
                    except Exception as e:
                        logger.warning(f"   Failed to parse embedding for row {idx}: {e}")
                
                results.append({
                    "id": row["ID"],
                    "prompt": row["Prompt"],
                    "response": row["Response"],
                    "embedding": embedding,
                    "usage_count": row.get("Usage_Count", 0)
                })
                logger.debug(f"   Loaded entry: {row['ID']} - {row['Prompt'][:50]}...")
            
            logger.info(f"✓ Loaded {len(results)} entries from {category} for similarity check")
            return results
        except Exception as e:
            logger.error(f"Error getting prompts: {e}")
            return []
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get statistics about the knowledge base"""
        stats = {
            "total_entries": 0,
            "by_category": {},
            "most_used": [],
            "recent_entries": []
        }
        
        try:
            xl = pd.ExcelFile(self.excel_path)
            
            for sheet_name in xl.sheet_names:
                df = pd.read_excel(xl, sheet_name=sheet_name)
                count = len(df)
                stats["total_entries"] += count
                stats["by_category"][sheet_name] = count
                
                if not df.empty and "Usage_Count" in df.columns:
                    top_used = df.nlargest(3, "Usage_Count")[["Prompt", "Usage_Count"]].to_dict("records")
                    stats["most_used"].extend(top_used)
            
            return stats
        except Exception as e:
            logger.error(f"Error getting statistics: {e}")
            return stats
